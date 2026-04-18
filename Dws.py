import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import queue
import os
import time
import pandas as pd
import numpy as np
from path_manager import get_initial_dir, load_path, save_path
from excel_utils import apply_excel_style

# ==================================================
# CORE EXCEL PIPELINE (NO GUI CODE HERE)
# ==================================================
def process_excel_pipeline(base_folder, append_file, jmsawb_file, cancel_flag, q):
    """
    ประมวลผล Excel ทั้งหมด
    ส่งสถานะกลับไปที่ GUI ผ่าน queue
    """

    def send(msg):
        q.put(msg)

    start_time = time.time()

    # -------------------------------
    # เตรียม Progress Calculation
    # -------------------------------
    base_files = [f for f in os.listdir(base_folder) if f.lower().endswith(".xlsx")]
    append_sheets = pd.ExcelFile(append_file, engine="openpyxl").sheet_names
    total_steps = len(base_files) + len(append_sheets) + 6
    step = 0

    def update_progress():
        nonlocal step
        step += 1
        percent = int(step / total_steps * 100)
        elapsed = time.time() - start_time
        eta = int((elapsed / step) * (total_steps - step)) if step else 0
        send(("PROGRESS", percent, eta))

    # -------------------------------
    # STEP 1: อ่านไฟล์ DWS 1–8
    # -------------------------------
    send("📂 อ่านไฟล์ DWS 1-8...")
    base_frames = []       
    
    for f in base_files:
        if cancel_flag.is_set():
            raise InterruptedError

        df = pd.read_excel(
            os.path.join(base_folder, f),
            usecols=[0, 2, 3, 4, 5],
            engine="openpyxl"
        )
        base_frames.append(df)
        update_progress()

    df_base = pd.concat(base_frames, ignore_index=True)

    # -------------------------------
    # STEP 2: อ่านไฟล์ DWS 9–11 (หลาย sheet)
    # -------------------------------
    send("📎 อ่านไฟล์ DWS 9-11...")
    append_frames = []

    for sheet in append_sheets:
        if cancel_flag.is_set():
            raise InterruptedError

        df = pd.read_excel(
            append_file,
            sheet_name=sheet,
            usecols=[3, 12, 13, 14, 15],
            engine="openpyxl"
        )
        append_frames.append(df)
        update_progress()

    df_append = pd.concat(append_frames, ignore_index=True)
    df_append.columns = df_base.columns

    # รวม base + append
    final_df = pd.concat([df_base, df_append], ignore_index=True)

    # -------------------------------
    # STEP 3: VLOOKUP จาก JMS-AWB
    # -------------------------------
    send("🔎 VLOOKUP...")
    jms = pd.read_excel(jmsawb_file, usecols=[0, 1], engine="openpyxl")
    jms.columns = ["KEY", "LOOK"]

    final_df.iloc[:, 0] = final_df.iloc[:, 0].astype(str)
    jms["KEY"] = jms["KEY"].astype(str)

    merged = final_df.merge(
        jms,
        left_on=final_df.columns[0],
        right_on="KEY",
        how="left"
    )

    final_df.insert(1, "DP ต้นทาง", merged["LOOK"])
    update_progress()

    # -------------------------------
    # STEP 4: สร้างคอลัมน์ จังหวัด (ตำแหน่ง 2–3)
    # -------------------------------
    send("🧮 สร้างคอลัมน์ จังหวัด...")

    def extract_province(val):
        if pd.isna(val):
            return np.nan
        s = str(val)
        return int(s[1:3]) if len(s) >= 3 and s[1:3].isdigit() else np.nan

    final_df.insert(6, "จังหวัด", final_df.iloc[:, 1].apply(extract_province))
    update_progress()

    # -------------------------------
    # STEP 5: ลบ Barcode ซ้ำ
    # -------------------------------
    send("🧾 ลบข้อมูลซ้ำ...")
    final_df = final_df.drop_duplicates(subset=final_df.columns[0])
    update_progress()

    # -------------------------------
    # STEP 6: กรองข้อมูลตาม Manual Rule
    # -------------------------------
    send("🧹 กรองข้อมูล...")

    # A: noread
    final_df = final_df[
        ~final_df.iloc[:, 0].astype(str).str.strip().str.lower().eq("noread")
    ]

    # B: ค่าว่าง
    final_df = final_df[
        final_df.iloc[:, 1].notna() &
        (final_df.iloc[:, 1].astype(str).str.strip() != "")
    ]

    # C: <= 3.00
    final_df = final_df[pd.to_numeric(final_df.iloc[:, 2], errors="coerce") <= 3.00]

    # D, E, F: <= 29.99
    for col in [3, 4, 5]:
        final_df = final_df[
            pd.to_numeric(final_df.iloc[:, col], errors="coerce") <= 29.99
        ]

    # G: >= 27
    final_df = final_df[
        pd.to_numeric(final_df.iloc[:, 6], errors="coerce") >= 27
    ]

    update_progress()

    # -------------------------------
    # FINISH
    # -------------------------------

    final_df.iloc[:, 0] = pd.to_numeric(final_df.iloc[:, 0], errors="coerce")

    send(("RESULT", final_df))
    send(f"✅ เสร็จสิ้น เหลือ {len(final_df)} แถว")
    send("__DONE__")
    
# ==================================================
# GUI APPLICATION
# ==================================================
class DwsTab(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.queue = queue.Queue()
        self.cancel_flag = threading.Event()
        self.result_df = None

        self.disable_widgets = []

        self.build_ui()
        self.after(100, self.listen_queue)


    # =======================
    # ENABLE / DISABLE UI
    # =======================
    def set_ui_state(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        for w in self.disable_widgets:
            try:
                w.config(state=state)
            except:
                pass

    # =======================
    # BUILD UI
    # =======================
    def build_ui(self):

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)   # Log ขยายได้

        FONT_NORMAL = ("Segoe UI", 10)
        FONT_TITLE = ("Segoe UI", 11, "bold")
        FONT_BUTTON = ("Segoe UI", 13, "bold")

        # ---------- Browse helpers ----------

        self.jms = tk.StringVar(value=load_path("jmsawb"))
        self.base = tk.StringVar(value=load_path("dws1_8"))
        self.append = tk.StringVar(value=load_path("dws9_11"))
        
        # ---------- Input Frame ----------
        frame_input = tk.LabelFrame(self, text=" Input Files ", font=FONT_TITLE)
        frame_input.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))


        rows = [
            ("JMS-AWB File", self.jms, self.browse_jms),
            ("DWS 1-8 Folder", self.base, self.browse_base),
            ("DWS 9-11 File", self.append, self.browse_append),
        ]


        for r, (lbl, var, cmd) in enumerate(rows):
            tk.Label(frame_input, text=lbl, font=FONT_NORMAL).grid(row=r, column=0, sticky="w", padx=10, pady=6)

            entry = tk.Entry(frame_input, textvariable=var, width=90, font=FONT_NORMAL)
            entry.grid(row=r, column=1)
            self.disable_widgets.append(entry)

            btn = tk.Button(
                frame_input,
                text="Browse",
                width=10,
                command=cmd
            )
            btn.grid(row=r, column=2, padx=5)
            self.disable_widgets.append(btn)

        # ---------- Progress ----------
        frm_progress = tk.LabelFrame(self, text=" Progress ", font=FONT_TITLE)
        frm_progress.grid(row=1, column=0, sticky="ew", padx=20, pady=10)

        frm_progress.grid_columnconfigure(0, weight=1)
        frm_progress.grid_columnconfigure(1, weight=0)

        self.progress = ttk.Progressbar(frm_progress, maximum=100)
        self.progress.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=20,
            pady=(10, 5)      
        )
        self.eta = tk.Label(
            frm_progress,
            text="ETA: --:--",
            font=FONT_NORMAL
        )
        self.eta.grid(
            row=1,
            column=0,
            sticky="e",
            padx=20,
            pady=(0, 10)
        )

        # ---------- Log ----------
        frm_log = tk.LabelFrame(
            self, text=" Log ",
            font=FONT_TITLE, padx=10, pady=6
        )
        frm_log.grid(row=2, column=0, sticky="nsew", padx=20, pady=(5, 10))

        frm_log.grid_rowconfigure(0, weight=1)
        frm_log.grid_columnconfigure(0, weight=1)

        scroll = tk.Scrollbar(frm_log)
        scroll.grid(row=0, column=1, sticky="ns")

        self.log = tk.Text(frm_log, font=("Consolas", 10), yscrollcommand=scroll.set)
        self.log.grid(row=0, column=0, sticky="nsew")

        scroll.config(command=self.log.yview)

        # ---------- Action Buttons ----------
        frm_action = tk.Frame(self)
        frm_action.grid(row=3, column=0, pady=(0, 20))

        self.btn_process = tk.Button(
            frm_action,
            text="▶  PROCESS",
            font=FONT_BUTTON,
            width=22,
            height=2,
            bg="#2ecc71",
            fg="white",
            command=self.start_process
        )
        self.disable_widgets.append(self.btn_process)

        self.btn_cancel = tk.Button(
            frm_action,
            text="✖  CANCEL",
            font=FONT_BUTTON,
            width=22,
            height=2,
            bg="#e74c3c",
            fg="white",
            command=self.cancel_process
        )
        self.btn_process.grid(in_=frm_action, row=0, column=0)
        self.btn_cancel.grid(in_=frm_action, row=0, column=0)
        self.btn_cancel.grid_remove()

    def browse_jms(self):
        init_dir = get_initial_dir("jmsawb", "last_dir")

        path = filedialog.askopenfilename(
            initialdir=init_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if path:
            self.jms.set(path)
            save_path("jmsawb", path)
            save_path("last_dir", os.path.dirname(path))


    def browse_base(self):
        init_dir = get_initial_dir("dws1_8", "last_dir")

        path = filedialog.askdirectory(initialdir=init_dir)

        if path:
            self.base.set(path)
            save_path("dws1_8", path)
            save_path("last_dir", path)


    def browse_append(self):
        init_dir = get_initial_dir("dws9_11", "last_dir")

        path = filedialog.askopenfilename(
            initialdir=init_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if path:
            self.append.set(path)
            save_path("dws9_11", path)
            save_path("last_dir", os.path.dirname(path))

    # =======================
    # START PROCESS
    # =======================
    def start_process(self):
        if not all([self.jms.get(), self.base.get(), self.append.get()]):
            messagebox.showwarning("Missing", "เลือก path ให้ครบ")
            return

        self.set_ui_state(False)
        self.config(cursor="watch")

        self.log.delete("1.0", tk.END)
        self.progress["value"] = 0
        self.eta.config(text="ETA: --:--")

        self.cancel_flag.clear()
        self.result_df = None

        self.btn_process.grid_remove()
        self.btn_cancel.grid()

        threading.Thread(target=self.worker, daemon=True).start()

    # =======================
    # BACKGROUND WORKER
    # =======================
    def worker(self):
        try:
            process_excel_pipeline(
                self.base.get(),
                self.append.get(),
                self.jms.get(),
                self.cancel_flag,
                self.queue
            )
        except Exception as e:
            self.queue.put(str(e))
            self.queue.put("__DONE__")

    # =======================
    # CANCEL PROCESS
    # =======================
    def cancel_process(self):
        self.cancel_flag.set()
        self.send("⛔ ผู้ใช้ยกเลิกการประมวลผล")

    # =======================
    # FINISH UI STATE
    # =======================
    def finish_ui(self):
        self.progress["value"] = 100
        self.btn_cancel.grid_remove()
        self.btn_process.grid()
        self.set_ui_state(True)
        self.config(cursor="")

    # =======================
    # SAVE RESULT
    # =======================
    def save_result(self):
        init_dir = get_initial_dir("output_dws", "last_dir")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=init_dir,
            initialfile="DwsQueryData.xlsx",
        )

        if not path:
            self.log.insert(tk.END, "⚠️ ยกเลิกการบันทึก\n")
            self.finish_ui()
            return

        try:
            self.result_df = self.result_df.rename(columns={
                self.result_df.columns[0]: "Barcode",
                self.result_df.columns[1]: "DP ต้นทาง",
                self.result_df.columns[2]: "Weight",
                self.result_df.columns[3]: "Length",
                self.result_df.columns[4]: "Width",
                self.result_df.columns[5]: "Height",
                self.result_df.columns[6]: "จังหวัด",
            })

            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                self.result_df.to_excel(writer, sheet_name="DATA", index=False)

            apply_excel_style(path, "DATA")

            from openpyxl import load_workbook

            wb = load_workbook(path)
            ws = wb["DATA"]

            for cell in ws["A"][1:]:
                cell.number_format = "0"

            wb.save(path)
    
            self.log.insert(tk.END, f"บันทึกไฟล์เรียบร้อย: {path}\n")
            messagebox.showinfo("เสร็จสิ้น", "บันทึกไฟล์เรียบร้อยแล้ว")
            save_path("output_dws", os.path.dirname(path))
            os.startfile(os.path.dirname(path))

        except Exception as e:
            messagebox.showerror("Save error", str(e))

        finally:
            self.finish_ui()

    # =======================
    # QUEUE LISTENER
    # =======================
    def listen_queue(self):
        try:
            while True:
                msg = self.queue.get_nowait()

                if msg == "__DONE__":
                    if self.result_df is not None:
                        self.save_result()
                    else:
                        self.finish_ui()

                elif isinstance(msg, tuple) and msg[0] == "PROGRESS":
                    _, percent, eta = msg
                    self.progress["value"] = percent
                    self.eta.config(text=f"ETA: {eta//60:02d}:{eta%60:02d}")

                elif isinstance(msg, tuple) and msg[0] == "RESULT":
                    self.result_df = msg[1]

                else:
                    self.log.insert(tk.END, msg + "\n")
                    self.log.see(tk.END)

        except queue.Empty:
            pass

        self.after(100, self.listen_queue)


