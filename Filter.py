import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
from path_manager import get_initial_dir, load_path, save_path
from openpyxl import load_workbook
from excel_utils import apply_excel_style
import configparser

# =======================
# CONSTANT
# =======================
#BLOCK_DP = {"829155", "829156", "829157", "829186"}


def is_excel_file(path):
    return path.lower().endswith((".xlsx", ".xls"))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "filter_rules.ini")


# =======================
# FILTER TAB
# =======================
class FilterTab(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.build_ui()
        self.load_block_list()
    # =======================
    # UI
    # =======================
    def build_ui(self):

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        FONT_NORMAL = ("Segoe UI", 10)
        FONT_TITLE = ("Segoe UI", 11, "bold")

        frame_input = tk.LabelFrame(self, text=" Input Files ", font=FONT_TITLE)
        frame_input.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))
        frame_input.grid_columnconfigure(1, weight=1)

        self.filter_auto = tk.StringVar(value=load_path("auto_filter"))
        self.filter_dws = tk.StringVar(value=load_path("dws_filter"))

        # Autopacking
        tk.Label(frame_input, text="Autopacking", font=FONT_NORMAL)\
            .grid(row=0, column=0, sticky="w", padx=10, pady=6)

        tk.Entry(frame_input, textvariable=self.filter_auto, width=90, font=FONT_NORMAL
        ).grid(row=0, column=1)

        tk.Button(
            frame_input, text="Browse", width=10,
            command=self.browse_auto
        ).grid(row=0, column=2, padx=5)

        # DWS
        tk.Label(frame_input, text="DWS", font=FONT_NORMAL)\
            .grid(row=1, column=0, sticky="w", padx=10, pady=6)

        tk.Entry(frame_input, textvariable=self.filter_dws, width=90, font=FONT_NORMAL
        ).grid(row=1, column=1)

        tk.Button(
            frame_input, text="Browse", width=10,
            command=self.browse_dws
        ).grid(row=1, column=2, padx=5)

        # ===== Middle Area (Rule + Progress) =====
        frame_middle = tk.Frame(self)
        frame_middle.grid(row=1, column=0, sticky="ew", padx=20, pady=(5, 5))

        # ===== Rule =====
        frame_rule = tk.LabelFrame(
            frame_middle,
            text="DP ที่ต้องกรองออก",
            font=FONT_TITLE
        )
        frame_rule.grid(row=0, column=0, sticky="nw", padx=(0, 20))

        frame_middle.grid_columnconfigure(0, weight=0)  # ซ้าย
        frame_middle.grid_columnconfigure(1, weight=1)  # ขวา

        self.dp_var = tk.StringVar()
        self.listbox = tk.Listbox(frame_rule, height=6, width=20)
        self.listbox.grid(row=0, column=0, rowspan=4, padx=5, pady=5)

        tk.Entry(frame_rule, textvariable=self.dp_var, width=15).grid(row=0, column=1, padx=5)
        tk.Button(frame_rule, text="Add", command=self.add_value).grid(row=1, column=1, sticky="ew", padx=5)
        tk.Button(frame_rule, text="Remove", command=self.remove_value).grid(row=2, column=1, sticky="ew", padx=5)

        # ===== Progress Panel =====
        style = ttk.Style()
        style.configure(
            "Big.Horizontal.TProgressbar",
            thickness=25   # ← ปรับเลขนี้ให้หนาตามต้องการ (20–30 กำลังสวย)
        )
        frame_process = tk.LabelFrame(
            frame_middle,
            text=" Progress ",
            font=FONT_TITLE
        )
        frame_process.grid(row=0, column=1, sticky="nsew")
        frame_process.grid_columnconfigure(0, weight=1)

        self.progress_var = tk.IntVar(value=0)

        self.progress_bar = ttk.Progressbar(
            frame_process,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            variable=self.progress_var,
            style="Big.Horizontal.TProgressbar"
        )
        self.progress_bar.grid(
            row=0, column=0,
            sticky="ew",
            padx=20,
            pady=(30, 10)
        )

        # Log
        frame_log = tk.LabelFrame(
            self, text=" Log ",
            font=FONT_TITLE, padx=10, pady=6
        )
        frame_log.grid(row=2, column=0, sticky="nsew", padx=20, pady=(5, 20))
        

        frame_log.grid_rowconfigure(0, weight=1)
        frame_log.grid_columnconfigure(0, weight=1)

        scroll = tk.Scrollbar(frame_log)
        scroll.grid(row=0, column=1, sticky="ns")

        self.log = tk.Text(frame_log, font=("Consolas", 10), yscrollcommand=scroll.set)
        self.log.grid(row=0, column=0, sticky="nsew")

        scroll.config(command=self.log.yview)

    # =======================
    # LOG
    # =======================
    def write_log(self, msg):
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)

    # =======================
    # BROWSE
    # =======================
    def browse_auto(self):
        init_dir = get_initial_dir("auto_filter", "last_dir")

        path = filedialog.askopenfilename(
            initialdir=init_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if path:
            self.filter_auto.set(path)
            save_path("auto_filter", path)
            save_path("last_dir", os.path.dirname(path))

            self.write_log("▶ เริ่มกรอง Autopacking อัตโนมัติ")
            self.process_autopacking()


    def browse_dws(self):
        init_dir = get_initial_dir("dws_filter", "last_dir")

        path = filedialog.askopenfilename(
            initialdir=init_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if path:
            self.filter_dws.set(path)
            save_path("dws_filter", path)
            save_path("last_dir", os.path.dirname(path))

            self.write_log("▶ เริ่มกรอง DWS อัตโนมัติ")
            self.process_dws()

    # ---------------- Rule Logic ----------------
    def add_value(self):
        val = self.dp_var.get().strip()
        if not val.isdigit():
            messagebox.showwarning("ผิดพลาด", "กรุณาใส่ตัวเลขเท่านั้น")
            return
        if val not in self.listbox.get(0, tk.END):
            self.listbox.insert(tk.END, val)
            self.dp_var.set("")
            self.save_block_list()

    def remove_value(self):
        sel = self.listbox.curselection()
        if sel:
            self.listbox.delete(sel[0])
            self.save_block_list()

    def load_block_list(self):
        cfg = configparser.ConfigParser()
        if os.path.exists(CONFIG_FILE):
            cfg.read(CONFIG_FILE, encoding="utf-8")

        values = cfg.get("FILTER", "block_dp", fallback="")
        for v in values.split(","):
            if v.strip():
                self.listbox.insert(tk.END, v.strip())

    def save_block_list(self):
        cfg = configparser.ConfigParser()
        cfg["FILTER"] = {
            "block_dp": ",".join(self.listbox.get(0, tk.END))
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            cfg.write(f)

    def get_block_list(self):
        return list(self.listbox.get(0, tk.END))
    
    def set_progress(self, value):
        self.progress_var.set(value)
        self.update_idletasks()

    # =======================
    # AUTOPACKING FILTER
    # =======================
    def process_autopacking(self):
        try:
            path = self.filter_auto.get()

            if not is_excel_file(path):
                raise ValueError("ไฟล์ Autopacking ไม่ใช่ Excel")

            self.write_log("📦 Autopacking: อ่านไฟล์...")
            self.set_progress(0)

            # STEP 1: Load file
            df = pd.read_excel(path)
            self.set_progress(10)

            if df.shape[1] < 4:
                raise ValueError("ไฟล์ Autopacking โครงสร้างไม่ถูกต้อง (ต้องมี ≥ 4 คอลัมน์)")

            before = len(df)

            # STEP 2: Load rule
            block = self.get_block_list()
            self.set_progress(25)

            # STEP 3: Filter DP
            df.iloc[:, 2] = df.iloc[:, 2].astype(str)
            df = df[~df.iloc[:, 2].isin(block)]
            self.set_progress(50)

            # STEP 4: Convert DP back
            df.iloc[:, 2] = pd.to_numeric(df.iloc[:, 2], errors="coerce")

            after = len(df)

            # แปลง Barcode เป็นตัวเลข ไม่มีทศนิยม
            df.iloc[:, 0] = (
                pd.to_numeric(df.iloc[:, 0], errors="coerce")
                .round(0)
                .astype("Int64")
            )
            self.set_progress(70)
            self.write_log(f"ก่อนกรอง: {before}")
            self.write_log(f"หลังกรอง: {after}")
            self.write_log(f"ถูกตัดออก: {before - after}")

            # STEP 5: Save
            self.save_file(df, "Autopacking_Filtered.xlsx")
            self.write_log("")
            self.set_progress(100)

        except Exception as e:
            self.write_log(f"❌ Autopacking Error: {e}")
            messagebox.showerror("Autopacking Error", str(e))

    # =======================
    # DWS FILTER
    # =======================
    def process_dws(self):
        try:
            path = self.filter_dws.get()

            if not is_excel_file(path):
                raise ValueError("ไฟล์ DWS ไม่ใช่ Excel")

            self.write_log("📊 DWS: อ่านไฟล์...")
            self.set_progress(0)

            # STEP 1: Load file
            df = pd.read_excel(path)
            self.set_progress(10)

            if df.shape[1] < 7:
                raise ValueError("ไฟล์ DWS โครงสร้างไม่ถูกต้อง (ต้องมี ≥ 7 คอลัมน์)")

            before = len(df)

            # STEP 2: Load rule
            block = self.get_block_list()
            self.set_progress(15)

            # STEP 3: Filter DP
            df.iloc[:, 1] = df.iloc[:, 1].astype(str)
            df = df[~df.iloc[:, 1].isin(block)]
            self.set_progress(30)

            # STEP 4: Convert DP back
            df.iloc[:, 1] = pd.to_numeric(df.iloc[:, 1], errors="coerce")
            self.set_progress(45)

            ## STEP 5: Convert numeric
            for col in [2, 3, 4, 5]:
                df.iloc[:, col] = pd.to_numeric(df.iloc[:, col], errors="coerce")
            self.set_progress(50)

            # STEP 6: Weight
            df = df[(df.iloc[:, 2] >= 1) & (df.iloc[:, 2] <= 2.99)]
            self.set_progress(55)

            # STEP 7: Dimension
            for col in [3, 4, 5]:
                df = df[(df.iloc[:, col] >= 1) & (df.iloc[:, col] <= 29)]

            after = len(df)
            self.set_progress(60)

            # แปลง Barcode เป็นตัวเลข ไม่มีทศนิยม
            df.iloc[:, 0] = (
                pd.to_numeric(df.iloc[:, 0], errors="coerce")
                .round(0)
                .astype("Int64")
            )
            self.set_progress(80)

            self.write_log(f"ก่อนกรอง: {before}")
            self.write_log(f"หลังกรอง: {after}")
            self.write_log(f"ถูกตัดออก: {before - after}")

            # STEP 8: Save
            self.save_file(df, "DWS_Filtered.xlsx")
            self.write_log("")
            self.set_progress(100)
        except Exception as e:
            self.write_log(f"❌ DWS Error: {e}")
            messagebox.showerror("DWS Error", str(e))

    # =======================
    # SAVE
    # =======================
    def save_file(self, df, default_name):
        init_dir = get_initial_dir("output_filter", "last_dir")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialdir=init_dir,
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )

        if not path:
            self.write_log("⚠️ ยกเลิกการบันทึก")
            return

        # เขียนไฟล์ด้วย pandas
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="DATA", index=False)

        # ===== บังคับ format cell =====
        wb = load_workbook(path)
        ws = wb["DATA"]

        for cell in ws["A"][1:]:  # ข้าม header
            cell.number_format = "0"

        # DP (Auto = C, DWS = B)
        # ใช้ชื่อ column ดีกว่า hardcode
        headers = [c.value for c in ws[1]]

        if "DP" in headers:
            dp_col = headers.index("DP") + 1  # Excel index
            col_letter = ws.cell(row=1, column=dp_col).column_letter

            for cell in ws[col_letter][1:]:
                cell.number_format = "0"

                
        wb.save(path)
        # ==============================

        apply_excel_style(path, "DATA")

        self.write_log(f"💾 บันทึกไฟล์: {path}")
        save_path("output_filter", os.path.dirname(path))
        os.startfile(os.path.dirname(path))
    
    '''def save_file(self, df, default_name):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")]
        )

        if not path:
            self.write_log("⚠️ ยกเลิกการบันทึก")
            return

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="DATA", index=False)

        self.write_log(f"💾 บันทึกไฟล์: {path}")
        os.startfile(os.path.dirname(path))'''
