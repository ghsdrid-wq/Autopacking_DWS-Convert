import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import pandas as pd
import os
import time
from path_manager import get_initial_dir, load_path, save_path
from excel_utils import apply_excel_style

# ==================================================
# Fonts
# ==================================================
FONT_NORMAL = ("Segoe UI", 10)
FONT_TITLE = ("Segoe UI", 11, "bold")
FONT_BUTTON = ("Segoe UI", 13, "bold")

# ==================================================
# Main Application
# ==================================================
class AutoTab(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.parent = parent
        self.stop_event = threading.Event()
        self.start_time = None
        self.result_df = None

        self.build_gui()

    # ==================================================
    # GUI
    # ==================================================
    def build_gui(self):

        self.grid_rowconfigure(2, weight=1)   # Log ขยายได้
        self.grid_columnconfigure(0, weight=1)

        # ---------- Input ----------
        frame_input = tk.LabelFrame(
            self, text=" Input Files ",
            font=FONT_TITLE, padx=15, pady=10
        )
        frame_input.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))
        
        self.jms_path = tk.StringVar(value=load_path("jmsawb"))
        self.billcode_path = tk.StringVar(value=load_path("billcode"))

        tk.Label(frame_input, text="JMS-AWB File", font=FONT_NORMAL)\
            .grid(row=0, column=0, sticky="w", pady=5)

        self.entry_jmsawb = tk.Entry(
            frame_input,
            font=FONT_NORMAL,
            textvariable=self.jms_path
        )
        self.entry_jmsawb.grid(row=0, column=1, sticky="ew", padx=10)

        self.btn_browse_jmsawb = tk.Button(
        frame_input, text="Browse",
        font=FONT_NORMAL, width=10,
        command=self.browse_jms
        )
        self.btn_browse_jmsawb.grid(row=0, column=2)

        tk.Label(frame_input, text="Billcode File", font=FONT_NORMAL)\
            .grid(row=1, column=0, sticky="w", pady=10)

        self.entry_billcode = tk.Entry(
            frame_input,
            font=FONT_NORMAL,
            textvariable=self.billcode_path
        )
        self.entry_billcode.grid(row=1, column=1, sticky="ew", padx=10)

        self.btn_browse_billcode = tk.Button(
            frame_input, text="Browse",
            font=FONT_NORMAL, width=10,
            command=self.browse_billcode
        )
        self.btn_browse_billcode.grid(row=1, column=2)

        frame_input.columnconfigure(1, weight=1)

        # ---------- Progress ----------
        '''frame_progress = tk.LabelFrame(
            self, text=" Progress ",
            font=FONT_TITLE, padx=15, pady=8
        )
        frame_progress.grid(row=1, column=0, sticky="ew", padx=20, pady=10)

        self.progress = ttk.Progressbar(
            frame_progress,
            orient="horizontal",
            mode="determinate"
        )
        self.progress.pack(fill="x")

        self.label_eta = tk.Label(
            frame_progress,
            text="ETA: --:--",
            font=FONT_NORMAL
        )
        self.label_eta.pack(anchor="w", pady=(5, 0))'''

        # ---------- Progress ----------
        style = ttk.Style()
        style.configure(
            "Big.Horizontal.TProgressbar",
            thickness=25   # ← ปรับเลขนี้ให้หนาตามต้องการ (20–30 กำลังสวย)
        )
        
        frame_progress = tk.LabelFrame(self, text=" Progress ", font=FONT_TITLE,)
        frame_progress.grid(row=1, column=0, sticky="ew", padx=20, pady=10)

        frame_progress.grid_columnconfigure(0, weight=1)
        frame_progress.grid_columnconfigure(1, weight=0)

        self.progress = ttk.Progressbar(frame_progress, maximum=100)
        self.progress.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=20,
            pady=(10, 5)      
        )
        self.eta = tk.Label(
            frame_progress,
            text="ETA: --:--",
            font=FONT_NORMAL
        )
        self.eta.grid(
            row=1,
            column=0,
            sticky="e",
            padx=20,
            pady=(0, 10)
        )

        # ---------- Log ----------
        frame_log = tk.LabelFrame(
            self, text=" Log ",
            font=FONT_TITLE, padx=10, pady=6
        )
        frame_log.grid(row=2, column=0, sticky="nsew", padx=20, pady=(5, 10))

        frame_log.grid_rowconfigure(0, weight=1)
        frame_log.grid_columnconfigure(0, weight=1)

        scroll = tk.Scrollbar(frame_log)
        scroll.grid(row=0, column=1, sticky="ns")

        self.text_log = tk.Text(
            frame_log,
            font=FONT_NORMAL,
            yscrollcommand=scroll.set
        )
        self.text_log.grid(row=0, column=0, sticky="nsew")

        scroll.config(command=self.text_log.yview)

        # ---------- Action (Footer) ----------
        frame_action = tk.Frame(self)
        frame_action.grid(row=3, column=0, pady=(0, 20))

        self.btn_process = tk.Button(
            frame_action,
            text="▶  PROCESS",
            font=FONT_BUTTON,
            width=22,
            height=2,
            bg="#2ecc71",
            fg="white",
            command=self.start_process
        )
        self.btn_process.grid(row=0, column=0)

        self.btn_cancel = tk.Button(
            frame_action,
            text="✖  CANCEL",
            font=FONT_BUTTON,
            width=22,
            height=2,
            bg="#e74c3c",
            fg="white",
            command=self.cancel_process
        )
        self.btn_cancel.grid(row=0, column=0)
        self.btn_cancel.grid_remove()

    # ==================================================
    # UI Helpers
    # ==================================================
    def log(self, message):
        self.after(0, lambda: (
            self.text_log.insert(tk.END, message + "\n"),
            self.text_log.see(tk.END)
        ))

    def set_progress(self, value):
        def _update():
            self.progress.config(value=value)
            if self.start_time and value > 5:
                elapsed = time.time() - self.start_time
                total_est = elapsed * 100 / value
                remain = max(0, total_est - elapsed)
                m, s = divmod(int(remain), 60)
                self.eta.config(text=f"ETA: {m:02d}:{s:02d}")
        self.after(0, _update)

    def show_process(self):
        self.btn_cancel.grid_remove()
        self.btn_process.grid()

    def show_cancel(self):
        self.btn_process.grid_remove()
        self.btn_cancel.grid()


    def disable_inputs(self):
        self.btn_browse_jmsawb.config(state="disabled")
        self.btn_browse_billcode.config(state="disabled")
        self.entry_jmsawb.config(state="disabled")
        self.entry_billcode.config(state="disabled")
        self.show_cancel()

    def enable_inputs(self):
        self.btn_browse_jmsawb.config(state="normal")
        self.btn_browse_billcode.config(state="normal")
        self.entry_jmsawb.config(state="normal")
        self.entry_billcode.config(state="normal")
        #self.label_eta.config(text="ETA: --:--")
        self.eta.config(text=f"ETA: --:--")
        self.show_process()

    # ==================================================
    # Browse
    # ==================================================
    def browse_jms(self):
        init_dir = get_initial_dir("jmsawb", "last_dir")

        path = filedialog.askopenfilename(
            initialdir=init_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if path:
            self.jms_path.set(path)
            save_path("jmsawb", path)
            save_path("last_dir", os.path.dirname(path))

    def browse_billcode(self):
        init_dir = get_initial_dir("billcode", "last_dir")

        path = filedialog.askopenfilename(
            initialdir=init_dir,
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )

        if path:
            self.billcode_path.set(path)
            save_path("billcode", path)
            save_path("last_dir", os.path.dirname(path))

    # ==================================================
    # Control
    # ==================================================
    def cancel_process(self):
        self.stop_event.set()
        self.log("⛔ ผู้ใช้ยกเลิกการประมวลผล")

    def start_process(self):
        if not self.entry_jmsawb.get() or not self.entry_billcode.get():
            messagebox.showwarning("Warning", "กรุณาเลือกไฟล์ให้ครบ")
            return

        self.stop_event.clear()
        self.start_time = time.time()
        self.progress["value"] = 0
        self.text_log.delete("1.0", tk.END)

        self.disable_inputs()
        threading.Thread(target=self.process_files, daemon=True).start()

    # ==================================================
    # Process Steps
    # ==================================================
    def step_load_billcode(self, path):
        self.log("📎 อ่านไฟล์ Billcode Detail...")
        xls = pd.ExcelFile(path, engine="xlrd")
        frames = []

        for sheet in xls.sheet_names:
            if self.stop_event.is_set():
                raise InterruptedError

            df = pd.read_excel(path, sheet_name=sheet, usecols=[0, 8], engine="xlrd")
            frames.append(df)

        df = pd.concat(frames, ignore_index=True)
        df.columns = ["Billcode", "Sort Code"]
        return df

    def step_clean_sortcode(self, df):
        self.log("🧹 ทำความสะอาด Sort Code...")
        cleaned = []
        total = len(df)

        for i, row in df.iterrows():
            if self.stop_event.is_set():
                raise InterruptedError

            sc = str(row["Sort Code"]).strip()[:2]
            if sc in ("", "na", "-", "S1"):
                continue

            row["Sort Code"] = sc
            cleaned.append(row)
            self.set_progress(int((i / total) * 40))

        return pd.DataFrame(cleaned)

    def step_vlookup_jms(self, df):
        self.log("🔎 VLOOKUP...")
        lookup = pd.read_excel(
            self.entry_jmsawb.get(),
            sheet_name="ส่งออกAWB",
            usecols=[0, 1]
        )
        lookup.columns = ["Billcode", "OriginCode"]

        df["Billcode"] = df["Billcode"].astype(str)
        lookup["Billcode"] = lookup["Billcode"].astype(str)

        df = df.merge(lookup, on="Billcode", how="left")
        df = df[df["OriginCode"].notna()]
        self.set_progress(60)
        return df

    def step_add_province(self, df):
        self.log("🧮 สร้างคอลัมน์ จังหวัด...")
        provinces = []
        total = len(df)

        for i, oc in enumerate(df["OriginCode"].astype(str)):
            if self.stop_event.is_set():
                raise InterruptedError

            provinces.append(int(oc[1:3]))
            self.set_progress(60 + int((i / total) * 20))

        df["จังหวัด"] = provinces
        return df

    def step_dedupe_and_filter(self, df):
        self.log("🧾 ลบข้อมูลซ้ำ...")
        df = df.drop_duplicates(subset=["Billcode"])

        self.log("🧹 กรองข้อมูล...")
        df = df[df["จังหวัด"] >= 27]

        return df.rename(columns={
            "Billcode": "Barcode",
            "Sort Code": "Destination Name",
            "OriginCode": "DP"
        })

    # ==================================================
    # Main Process
    # ==================================================
    def process_files(self):
        try:
            df = self.step_load_billcode(self.entry_billcode.get())
            df = self.step_clean_sortcode(df)
            df = self.step_vlookup_jms(df)
            df = self.step_add_province(df)
            df = self.step_dedupe_and_filter(df)

            df["Barcode"] = pd.to_numeric(df["Barcode"], errors="coerce")

            self.result_df = df
            self.set_progress(90)

            self.log(f"✅ เสร็จสิ้น เหลือ {len(df)} แถว")
            self.after(0, self.on_process_finished)

        except InterruptedError:
            self.log("⛔ การประมวลผลถูกยกเลิก")
            self.after(0, self.enable_inputs)

        except Exception as e:
            self.log(f"❌ Error: {e}")
            self.after(0, lambda: messagebox.showerror("Error", str(e)))
            self.after(0, self.enable_inputs)

    # ==================================================
    # Save File
    # ==================================================
    def on_process_finished(self):
        init_dir = get_initial_dir("output_auto", "last_dir")
        output_file = filedialog.asksaveasfilename(
            title="บันทึกไฟล์ผลลัพธ์",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialdir=init_dir,
            initialfile="AutopackingData.xlsx",
        )


        if not output_file:
            self.enable_inputs()
            return

        threading.Thread(
            target=self.save_excel_file,
            args=(output_file,),
            daemon=True
        ).start()

    def save_excel_file(self, output_file):
        try:
            self.log("💾 กำลังบันทึกไฟล์ Excel...")
            self.set_progress(95)

            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                self.result_df.to_excel(writer, sheet_name="DATA", index=False)

            apply_excel_style(output_file, "DATA")

            from openpyxl import load_workbook

            wb = load_workbook(output_file)
            ws = wb["DATA"]

            # บังคับ Barcode เป็น Number
            for cell in ws["A"][1:]:
                cell.number_format = "0"

            wb.save(output_file)

            self.set_progress(100)
            self.after(0, lambda: (
                #self.label_eta.config(text="ETA: 00:00")
                self.eta.config(text=f"ETA: --:--"),
                self.log(f"บันทึกไฟล์เรียบร้อย: {output_file}"),
                messagebox.showinfo("Success", "ประมวลผลเสร็จเรียบร้อย"),
                save_path("output_auto", os.path.dirname(output_file)),
                os.startfile(os.path.dirname(output_file)),
                self.enable_inputs()
            ))

        except Exception as e:
            self.after(0, lambda: (
                self.log(f"❌ Error ตอนบันทึกไฟล์: {e}"),
                messagebox.showerror("Error", str(e)),
                self.enable_inputs()
            ))


